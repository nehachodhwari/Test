from ..app import db
from automate.models.models import AdUsers, ProfileMapping, ExtensionIdentification, LocationMapping
import csv
from ..deprovision.cucm_utils import get_user_data, get_phone_data
from ..deprovision.utils import get_user_configurations
from ..deprovision.usecase import deprovision_any_user
from datetime import datetime
import openpyxl
from sqlalchemy import exc
import time as ti

ALLOWED_EXTENSIONS = {'csv'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def read_file(filename):
    with open('automate/Uploadfiles/' + filename) as f:
        file_reader = csv.DictReader(f)
        users = []
        for row in file_reader:
            users.append(dict(row))
        return users


def deprovision_users(user_list):
    users_deprovisioned = []
    users_not_deprovisioned = []
    error_messages = []
    for user in user_list:
        get_user = find_user(user['USERID'])
        print(get_user)
        if get_user:
            user_data = get_User_Details(user['USERID'])
            user_basic_details = get_user_configurations(user['USERID'])
            output = user_data.copy()
            if user_basic_details:
                output.update(user_basic_details)
                deprovision_user = deprovision_any_user(**output)
                if deprovision_user[0]:
                    users_deprovisioned.append(user['USERID'])
                else:
                    users_not_deprovisioned.append(user['USERID'])
            else:
                error_messages.append('Details not found for user {user}'.format(user=user['USERID']))
        else:
            error_messages.append("User {user} Not found.".format(user=user['USERID']))
    data_sheet_name = create_Data_Sheet(users_deprovisioned, users_not_deprovisioned)
    print(data_sheet_name)
    return data_sheet_name, error_messages



def find_user(userid):
    """
	@param userid: check if user exists
	@return: boolean
	"""

    user = None
    retry_flag = True
    retry_count = 0
    while retry_flag and retry_count < 5:
        try:
            user = db.session.query(AdUsers.samAccountName).filter(
                AdUsers.samAccountName.ilike(userid)).all()
            retry_flag = False
        except exc.SQLAlchemyError:
            retry_count = retry_count + 1
            ti.sleep(1)
        finally:
            db.session.close()
    if user:
        return True
    else:
        return False


def get_User_Details(user_id):
    """
    Function for finding user details to deprovision
    """

    user_data = get_user_data(user_id)
    print(user_data)
    user_basic_details = get_user_configurations(user_id)
    print(user_basic_details)
    output = user_data.copy()
    if user_basic_details:
        output.update(user_basic_details)
    return output


def create_Data_Sheet(list1,list2):
    """
    Function to create excel sheet for users
    provisioned and not provisioned during bulk provision task
    """
    work_book = openpyxl.Workbook()
    data_sheet = work_book.active
    now = datetime.now()
    data_sheet_title = "deprov"+now.strftime("%d%m%Y%H%M%S")
    c1 = data_sheet.cell(row=1, column=1)
    c2 = data_sheet.cell(row=1, column=2)
    c1.value = "Deprovisioned"
    c2.value = "Not deprovisioned"
    column, row = 1, 2
    for user in list1:
        sheet_cell = data_sheet.cell(row, column)
        sheet_cell.value = user
        row = row + 1
    column, row = 2, 2
    for user in list2:
        sheet_cell = data_sheet.cell(row, column)
        sheet_cell.value = user
        row = row + 1
    work_book.save("automate/static/"+data_sheet_title+".xlsx")
    return(data_sheet_title)