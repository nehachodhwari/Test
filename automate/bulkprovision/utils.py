import csv

from flask import jsonify
import openpyxl
from datetime import datetime
from sqlalchemy import exc
from urllib3 import request
from ..app import db
from automate.models.models import AdUsers, ProfileMapping, ExtensionIdentification, LocationMapping
from ..quickprovision.utils import list_line, check_valid_extension, get_new_line, get_form_data
from ..quickprovision.config import PHONE_BUTTON_TEMPLATE
from ..quickprovision.quickusecase_functions import quick_provision_sales, quick_provision_account_manager\
    , quick_provision_executive_profile,quick_provision_knowledge_worker
import time as ti

ALLOWED_EXTENSIONS = {'csv'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def read_file(filename):
    with open('automate/Uploadfiles/' + filename) as f:
        file_reader = csv.DictReader(f)
        users = []
        for row in file_reader:
            users.append(dict(row))
        return users


def provision_users(user_list):
    users_provisioned = []
    users_not_provision = []
    error_messages = []
    for user in user_list:
        get_user = find_user(user['USERID'])
        print(get_user)
        if get_user:
            user_detail = get_user_details(user['USERID'])
            print(type(user_detail))
            if user_detail[1]:
                user_data = get_user_configuration(user_detail[0], user)
                user_provision = provision(user_data)
                if user_provision == 'Success':
                    users_provisioned.append(user['USERID'])
                else:
                    users_not_provision.append(user['USERID'])
            else:
                error_messages.append('Details not found for user {user}'.format(user=user['USERID']))
        else:
            error_messages.append( "User {user} Not found.".format(user=user['USERID']))
    data_sheet_name = create_Data_Sheet(users_provisioned,users_not_provision)
    print(data_sheet_name)
    return data_sheet_name, error_messages





def find_user(userid):
	"""
	@param userid: check if user exists
	@return: boolean
	"""

	user = None
	retry_flag = True
	retry_count = 0
	while retry_flag and retry_count < 5:
		try:
			user = db.session.query(AdUsers.samAccountName).filter(
				AdUsers.samAccountName.ilike(userid)).all()
			retry_flag = False
		except exc.SQLAlchemyError:
			retry_count = retry_count + 1
			ti.sleep(1)
		finally:
			db.session.close()
	if user:
		return True
	else:
		return False




def get_user_details(user_id):
    retry_flag = True
    retry_count = 0
    user_data = None
    while retry_flag and retry_count < 5:
        try:
            user_data = AdUsers.query.filter(AdUsers.samAccountName == user_id).all()
            retry_flag = False
        except exc.SQLAlchemyError:
            retry_count = retry_count + 1
            ti.sleep(1)
        finally:
            db.session.close()
    if user_data:
        for data in user_data:
            try:
                profile_name = db.session.query(ProfileMapping).filter(data.Title == ProfileMapping.Title).\
                    first().ProfileName
                extension_pick = db.session.query(ExtensionIdentification.Extensionlocation).\
                    first().Extensionlocation
                location = data.Location.strip()
                mapped_location = db.session.query(LocationMapping.SiteCode).filter(
                    LocationMapping.Location == location).first().SiteCode
                extension = data.TelephoneNumber if extension_pick == 'TelephoneNumber' else \
                    (data.WorkphoneNumber if extension_pick == 'WorkPhone' else '')
                if not check_valid_extension(extension, mapped_location):
                    extension_list = list_line(mapped_location)
                    print(extension_list)
                    extension = get_new_line(extension_list[0], mapped_location)
                output = {'fname': data.FirstName.strip(), 'lname': data.LastName.strip(),
                          'email': data.Email.strip(), 'location':mapped_location, 'extension': extension,
                          'title': profile_name}
                print(output)
                return output, True
            except exc.SQLAlchemyError:
                return {'message':'Internal error occurred'}, False, 500
    return {'message': 'No User data found for the user'}, False, 404


def get_user_configuration(user_details, user):
    print("User Details:    ", user_details)
    mac = user['MAC']
    if mac == 'NA' or 'None':
        mac = None
    destination = user['DESTINATION']
    if destination == 'NA' or 'None':
        destination = None
    product = user['PRODUCT']
    protocol = user['PROTOCOL']
    if user_details['title'] == "Knowledge Worker":
        phone_button_template = None
    else:
        if product == 'NA' or 'None':
            product = "Cisco 7962"
            protocol = 'SIP'
        phone_button_template = PHONE_BUTTON_TEMPLATE[product][protocol]["Default"]
    output = {'user_profile': user_details['title'],
              'phonebuttontemplate': phone_button_template,
              'calling_privilege': 'international',
              'Mac': mac, 'destination': destination,
              'location': user_details['location'],
              'user_id': user['USERID'], 'product': product,
              'protocol': protocol, 'extension': user_details['extension'],
              'first_name': user_details['fname'],
              'last_name': user_details['lname'],
              'email': user_details['email'],
              'alerting_name': user_details['fname'] + ' ' + user_details['lname']}
    return output


def provision(user_data):
    user_profile = user_data['user_profile']
    if user_profile == 'Executive':
        result = quick_provision_executive_profile(**user_data)
    elif user_profile == 'Knowledge Worker':
        result = quick_provision_knowledge_worker(**user_data)
    elif user_profile == 'Sales':
        result = quick_provision_sales(**user_data)
    elif user_profile == 'Account Manager':
        result = quick_provision_account_manager(**user_data)
    print(result)
    return result[1]

def create_Data_Sheet(list1,list2):
    """
    Function to create excel sheet for users
    provisioned and not provisioned during bulk provision task
    """
    work_book = openpyxl.Workbook()
    data_sheet = work_book.active
    now = datetime.now()
    data_sheet_title = now.strftime("%d%m%Y%H%M%S")
    c1 = data_sheet.cell(row=1, column=1)
    c2 = data_sheet.cell(row=1, column=2)
    c1.value = "Provisioned"
    c2.value = "Not Provisioned"
    column, row = 1, 2
    for user in list1:
        sheet_cell = data_sheet.cell(row, column)
        sheet_cell.value = user
        row = row + 1
    column, row = 2, 2
    for user in list2:
        sheet_cell = data_sheet.cell(row, column)
        sheet_cell.value = user
        row = row + 1
    work_book.save("automate/static/"+data_sheet_title+".xlsx")
    return(data_sheet_title)

